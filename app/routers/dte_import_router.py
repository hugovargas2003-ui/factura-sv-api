"""
DTE Import Router — Importar DTEs históricos emitidos en otros sistemas.
Los DTEs importados tienen estado 'IMPORTADO' — no se transmiten al MH.
Se incluyen en reportes (F-07, Libro Ventas, etc.)
"""
import logging
import io
import csv as csv_mod
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/dtes", tags=["dtes-import"])

from app.dependencies import get_current_user, get_supabase


@router.post("/import-historico")
async def import_dtes_historicos(
    file: UploadFile = File(...),
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    """
    Import DTEs emitidos en otro sistema (ContaPortable, EasyFact, etc.)
    NO firma, NO transmite al MH, NO consume créditos.
    Estado: 'IMPORTADO' — se incluye en reportes pero no en re-transmisión.
    """
    org_id = user.get("org_id")
    content = await file.read()
    filename = file.filename or ""

    rows: list[dict] = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv_mod.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl requerido para XLSX")
    else:
        raise HTTPException(status_code=400, detail="Use CSV o XLSX")

    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacío")

    col_aliases = {
        "fecha": ["fecha", "date", "fec_emi", "fecha_emision", "fecha emision"],
        "tipo_dte": ["tipo", "tipo_dte", "tipo dte", "tipo_documento", "clase_documento", "clase"],
        "numero_control": ["numero_control", "numero control", "no_control", "correlativo", "control"],
        "codigo_generacion": ["codigo_generacion", "codigo generacion", "codigo", "uuid", "cod_generacion"],
        "receptor_nit": ["nit_receptor", "receptor_nit", "nit_cliente", "nit cliente", "nit"],
        "receptor_nombre": ["nombre_receptor", "receptor_nombre", "cliente", "nombre cliente", "razon_social"],
        "total_gravada": ["gravada", "total_gravada", "venta_gravada", "gravadas", "ventas_gravadas"],
        "total_exenta": ["exenta", "total_exenta", "venta_exenta", "exentas"],
        "total_no_suj": ["no_sujeta", "no_suj", "total_no_suj", "ventas_no_sujetas"],
        "iva": ["iva", "debito_fiscal", "iva_debito", "impuesto"],
        "monto_total": ["total", "monto_total", "total_pagar", "monto", "valor"],
        "sello_recepcion": ["sello", "sello_recepcion", "sello_mh"],
        "condicion_pago": ["condicion", "condicion_pago", "forma_pago", "pago"],
    }

    def find_val(row, aliases):
        for alias in aliases:
            for key in row:
                if str(key).strip().lower() == alias:
                    v = row[key]
                    return str(v).strip() if v is not None else None
        return None

    def safe_float(val):
        if val is None:
            return 0.0
        try:
            return float(str(val).replace(",", "").replace("$", "").strip())
        except Exception:
            return 0.0

    created = 0
    skipped = 0
    errors = []
    warnings = []

    for i, row in enumerate(rows):
        try:
            receptor_nombre = find_val(row, col_aliases["receptor_nombre"])
            fecha = find_val(row, col_aliases["fecha"])

            if not fecha:
                errors.append(f"Fila {i+2}: sin fecha")
                continue

            tipo_dte = find_val(row, col_aliases["tipo_dte"]) or "01"
            tipo_dte = str(tipo_dte).zfill(2)

            total_gravada = safe_float(find_val(row, col_aliases["total_gravada"]))
            total_exenta = safe_float(find_val(row, col_aliases["total_exenta"]))
            total_no_suj = safe_float(find_val(row, col_aliases["total_no_suj"]))
            iva = safe_float(find_val(row, col_aliases["iva"]))
            monto_total = safe_float(find_val(row, col_aliases["monto_total"]))

            # Auto-calculate missing fields
            if iva == 0 and total_gravada > 0:
                if tipo_dte == "03":  # CCF: IVA adicional
                    iva = round(total_gravada * 0.13, 2)
                elif tipo_dte == "01":  # Factura: IVA incluido
                    iva = round(total_gravada - total_gravada / 1.13, 2)

            if monto_total == 0:
                if tipo_dte == "03":
                    monto_total = total_gravada + iva
                else:
                    monto_total = total_gravada + total_exenta + total_no_suj

            codigo_gen = find_val(row, col_aliases["codigo_generacion"])
            if not codigo_gen:
                codigo_gen = f"IMP-{uuid.uuid4()}"

            numero_ctrl = find_val(row, col_aliases["numero_control"]) or ""
            sello = find_val(row, col_aliases["sello_recepcion"])
            receptor_nit = find_val(row, col_aliases["receptor_nit"]) or ""

            # Check duplicate by codigo_generacion
            existing = supabase.table("dtes") \
                .select("id").eq("org_id", org_id).eq("codigo_generacion", codigo_gen).execute()

            if existing.data:
                skipped += 1
                continue

            # Insert with estado='IMPORTADO'
            record = {
                "org_id": org_id,
                "tipo_dte": tipo_dte,
                "numero_control": numero_ctrl,
                "codigo_generacion": codigo_gen,
                "sello_recibido": sello,
                "estado": "IMPORTADO",
                "fecha_emision": fecha,
                "monto_total": monto_total,
                "receptor_nombre": receptor_nombre or "Sin receptor",
                "receptor_nit": receptor_nit,
                "json_data": {
                    "source": "import_historico",
                    "row": i + 2,
                    "resumen": {
                        "totalGravada": total_gravada,
                        "totalExenta": total_exenta,
                        "totalNoSuj": total_no_suj,
                        "subTotalVentas": total_gravada + total_exenta + total_no_suj,
                        "montoTotalOperacion": monto_total,
                        "totalPagar": monto_total,
                    },
                },
            }

            supabase.table("dtes").insert(record).execute()
            created += 1

        except Exception as e:
            errors.append(f"Fila {i+2}: {str(e)}")

    return {
        "created": created,
        "skipped_duplicates": skipped,
        "errors": errors,
        "warnings": warnings,
        "total_rows": len(rows),
        "note": "DTEs importados con estado 'IMPORTADO'. Se incluyen en reportes pero no se transmiten al MH.",
    }


@router.get("/import-historico/template")
async def download_template(
    tipo: str = Query("01", description="01=Factura, 03=CCF, 14=Sujeto Excluido"),
    user=Depends(get_current_user),
):
    """Descargar plantilla CSV para import de DTEs históricos."""
    from fastapi.responses import Response

    if tipo == "03":
        header = "fecha,tipo_dte,numero_control,codigo_generacion,sello_recepcion,receptor_nit,receptor_nombre,total_gravada,iva_debito,monto_total,condicion_pago"
        sample = "2026-01-15,03,DTE-03-XXXXX-000001234,UUID-AQUI,SELLO-AQUI,0614-XXXXXX-XXX-X,Empresa ABC S.A. de C.V.,1000.00,130.00,1130.00,1"
    elif tipo == "14":
        header = "fecha,tipo_dte,numero_control,receptor_nombre,total_gravada,monto_total"
        sample = "2026-01-15,14,DTE-14-XXXXX-000001234,Juan Pérez,500.00,500.00"
    else:
        header = "fecha,tipo_dte,numero_control,receptor_nombre,monto_total"
        sample = "2026-01-15,01,DTE-01-XXXXX-000001234,Consumidor Final,56.50"

    csv_content = f"{header}\n{sample}\n"

    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=plantilla_import_tipo{tipo}.csv"},
    )
