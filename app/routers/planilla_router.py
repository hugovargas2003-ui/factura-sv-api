"""
Planilla Router — Gestión de Nómina Mensual
8 endpoints para upload, cálculo automático y generación de líneas F14.
NO modifica ningún archivo existente.
"""
import logging
import io
import csv
from typing import Optional
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/planilla", tags=["planilla"])

from app.dependencies import get_current_user, get_supabase


# ═══════════════════════════════════════════
# MODELOS
# ═══════════════════════════════════════════

class EmpleadoManual(BaseModel):
    nombre: str
    nit: Optional[str] = None
    dui: Optional[str] = None
    codigo_ingreso: str = "01"
    salario_base: float
    bonificaciones: float = 0
    horas_extras: float = 0
    comisiones: float = 0
    isr_override: Optional[float] = None

class EmpleadoUpdate(BaseModel):
    nombre: Optional[str] = None
    salario_base: Optional[float] = None
    bonificaciones: Optional[float] = None
    isr_retenido: Optional[float] = None
    isr_override_motivo: Optional[str] = None
    afp: Optional[float] = None
    isss: Optional[float] = None


# ═══════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════

def _calcular_empleado(emp: dict, supabase) -> dict:
    """Calcula todas las deducciones de un empleado."""
    salario = float(emp.get("salario_base", 0))
    bonif = float(emp.get("bonificaciones", 0))
    hextras = float(emp.get("horas_extras", 0))
    comisiones = float(emp.get("comisiones", 0))

    total_devengado = salario + bonif + hextras + comisiones

    afp_emp = round(salario * 0.0725, 2)
    isss_emp = round(min(salario, 1000) * 0.03, 2)

    base_imponible = salario - afp_emp - isss_emp
    isr = 0.0
    if base_imponible > 0:
        tramos = supabase.table("renta_tabla_isr") \
            .select("*").eq("vigencia_desde", "2024-01-01").order("tramo").execute()
        for t in (tramos.data or []):
            desde = float(t["desde"])
            hasta = float(t["hasta"]) if t["hasta"] else float('inf')
            if desde <= base_imponible <= hasta:
                tasa = float(t["tasa"])
                cuota = float(t["cuota_fija"])
                exceso = float(t["sobre_exceso"])
                if tasa > 0:
                    isr = round(cuota + (base_imponible - exceso) * tasa, 2)
                break

    if emp.get("isr_override") is not None:
        isr = float(emp["isr_override"])

    salario_neto = round(total_devengado - afp_emp - isss_emp - isr, 2)

    afp_pat = round(salario * 0.0875, 2)
    isss_pat = round(min(salario, 1000) * 0.075, 2)
    insaforp = round(salario * 0.01, 2)

    return {
        "total_devengado": total_devengado,
        "afp": afp_emp,
        "isss": isss_emp,
        "isr_retenido": isr,
        "salario_neto": salario_neto,
        "afp_patronal": afp_pat,
        "isss_patronal": isss_pat,
        "insaforp": insaforp,
    }


def _validate_planilla_data(rows_parsed: list) -> dict:
    """
    Validación post-parse de datos de planilla.
    Retorna {errors: [...], warnings: [...]}
    errors = bloquean el upload
    warnings = se muestran pero no bloquean
    """
    errors = []
    warnings = []
    seen_ids: set[str] = set()

    SALARIO_MINIMO_SV = 365.00  # Comercio/servicios 2024
    SALARIO_MAXIMO_RAZONABLE = 50000.00

    for i, emp in enumerate(rows_parsed):
        row_num = i + 2  # +2 por header + 0-indexed
        nombre = str(emp.get("nombre", "")).strip()
        salario = emp.get("salario_base", 0)
        nit = str(emp.get("nit", "") or "")
        dui = str(emp.get("dui", "") or "")

        # Errores (bloquean)
        if salario is not None and float(salario) < 0:
            errors.append(f"Fila {row_num}: Salario negativo (${salario}) para {nombre}")

        if not nombre:
            errors.append(f"Fila {row_num}: Nombre vacío")
            continue

        # Warnings (no bloquean)
        if salario is not None and float(salario) < SALARIO_MINIMO_SV and float(salario) > 0:
            warnings.append(f"Fila {row_num}: Salario ${salario} menor al mínimo (${SALARIO_MINIMO_SV}) para {nombre}")

        if salario is not None and float(salario) > SALARIO_MAXIMO_RAZONABLE:
            warnings.append(f"Fila {row_num}: Salario ${salario} inusualmente alto para {nombre}")

        # Duplicados por NIT o DUI
        id_key = nit or dui or nombre.lower()
        if id_key in seen_ids:
            warnings.append(f"Fila {row_num}: Posible duplicado — {nombre} ({nit or dui})")
        seen_ids.add(id_key)

        # NIT formato (14 dígitos)
        if nit and len(nit.replace("-", "").replace(" ", "")) not in (0, 14):
            warnings.append(f"Fila {row_num}: NIT '{nit}' no tiene 14 dígitos para {nombre}")

        # DUI formato (9 dígitos + guión + 1)
        if dui:
            dui_clean = dui.replace("-", "").replace(" ", "")
            if len(dui_clean) not in (0, 9, 10):
                warnings.append(f"Fila {row_num}: DUI '{dui}' formato inválido para {nombre}")

    return {"errors": errors, "warnings": warnings}


def _recalc_planilla_totals(supabase, planilla_id: str):
    """Recalcula totales del resumen de planilla."""
    emps = supabase.table("planilla_empleados") \
        .select("*").eq("planilla_id", planilla_id).execute()

    data = emps.data or []
    totals = {
        "total_empleados": len(data),
        "total_salarios": sum(float(e.get("total_devengado", 0)) for e in data),
        "total_afp_emp": sum(float(e.get("afp", 0)) for e in data),
        "total_afp_pat": sum(float(e.get("afp_patronal", 0)) for e in data),
        "total_isss_emp": sum(float(e.get("isss", 0)) for e in data),
        "total_isss_pat": sum(float(e.get("isss_patronal", 0)) for e in data),
        "total_insaforp": sum(float(e.get("insaforp", 0)) for e in data),
        "total_isr": sum(float(e.get("isr_retenido", 0)) for e in data),
        "total_neto": sum(float(e.get("salario_neto", 0)) for e in data),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    supabase.table("planilla_resumen").update(totals).eq("id", planilla_id).execute()


# ═══════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════

@router.get("")
async def listar_planillas(
    periodo: Optional[str] = None,
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    org_id = user.get("org_id")
    query = supabase.table("planilla_resumen").select("*").eq("org_id", org_id)
    if periodo:
        query = query.eq("periodo", periodo)
    result = query.order("created_at", desc=True).execute()
    return {"planillas": result.data or []}


@router.post("/upload")
async def upload_planilla(
    periodo: str = Query(...),
    file: UploadFile = File(...),
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    """Upload CSV/XLSX de planilla. Calcula automáticamente AFP, ISSS, ISR."""
    org_id = user.get("org_id")

    content = await file.read()
    filename = file.filename or ""

    rows = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl required for XLSX files")
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o XLSX.")

    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacío o sin datos")

    col_map = {
        "nombre": ["nombre", "empleado", "name", "nombres", "apellidos", "nombre completo"],
        "dui": ["dui", "documento", "identidad"],
        "nit": ["nit", "nit/nif"],
        "salario_base": ["salario", "salario_base", "sueldo", "salario base", "monto", "ingreso"],
        "bonificaciones": ["bonificaciones", "bonificacion", "bono", "gratificacion"],
        "horas_extras": ["horas_extras", "horas extras", "overtime"],
        "comisiones": ["comisiones", "comision"],
    }

    def find_col(row, aliases):
        for alias in aliases:
            for key in row:
                if str(key).strip().lower() == alias:
                    return row[key]
        return None

    # Pre-validate parsed data
    pre_parsed = []
    for row in rows:
        pre_parsed.append({
            "nombre": str(find_col(row, col_map["nombre"]) or ""),
            "salario_base": float(str(find_col(row, col_map["salario_base"]) or 0).replace(",", "").replace("$", "").strip() or 0),
            "nit": str(find_col(row, col_map["nit"]) or ""),
            "dui": str(find_col(row, col_map["dui"]) or ""),
        })
    validation = _validate_planilla_data(pre_parsed)
    if validation["errors"]:
        return {
            "planilla_id": None,
            "empleados_created": 0,
            "errors": validation["errors"],
            "warnings": validation["warnings"],
            "status": "validation_failed",
        }

    planilla = supabase.table("planilla_resumen").insert({
        "org_id": org_id,
        "periodo": periodo,
        "nombre": f"Planilla {periodo}",
        "status": "draft",
    }).execute()
    planilla_id = planilla.data[0]["id"]

    empleados_created = 0
    errors = []

    for i, row in enumerate(rows):
        try:
            nombre = find_col(row, col_map["nombre"])
            if not nombre:
                errors.append(f"Fila {i+2}: sin nombre")
                continue

            salario = find_col(row, col_map["salario_base"])
            if not salario:
                errors.append(f"Fila {i+2}: sin salario para {nombre}")
                continue

            salario = float(str(salario).replace(",", "").replace("$", "").strip())
            bonif = float(str(find_col(row, col_map["bonificaciones"]) or 0).replace(",", ""))
            hextras = float(str(find_col(row, col_map["horas_extras"]) or 0).replace(",", ""))
            comis = float(str(find_col(row, col_map["comisiones"]) or 0).replace(",", ""))

            emp_data = {
                "salario_base": salario,
                "bonificaciones": bonif,
                "horas_extras": hextras,
                "comisiones": comis,
            }

            calc = _calcular_empleado(emp_data, supabase)

            supabase.table("planilla_empleados").insert({
                "planilla_id": planilla_id,
                "org_id": org_id,
                "nombre": str(nombre).strip(),
                "nit": str(find_col(row, col_map["nit"]) or "").strip() or None,
                "dui": str(find_col(row, col_map["dui"]) or "").strip() or None,
                "codigo_ingreso": "01",
                "salario_base": salario,
                "bonificaciones": bonif,
                "horas_extras": hextras,
                "comisiones": comis,
                **calc,
            }).execute()
            empleados_created += 1

        except Exception as e:
            errors.append(f"Fila {i+2}: {str(e)}")

    _recalc_planilla_totals(supabase, planilla_id)

    return {
        "planilla_id": planilla_id,
        "empleados_created": empleados_created,
        "errors": errors,
        "warnings": validation.get("warnings", []),
        "status": "draft",
    }


@router.get("/{planilla_id}")
async def detalle_planilla(
    planilla_id: str,
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    org_id = user.get("org_id")

    planilla = supabase.table("planilla_resumen") \
        .select("*").eq("id", planilla_id).eq("org_id", org_id).execute()
    if not planilla.data:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")

    empleados = supabase.table("planilla_empleados") \
        .select("*").eq("planilla_id", planilla_id).order("nombre").execute()

    return {"planilla": planilla.data[0], "empleados": empleados.data or []}


@router.put("/{planilla_id}/empleados/{empleado_id}")
async def actualizar_empleado(
    planilla_id: str,
    empleado_id: str,
    body: EmpleadoUpdate,
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    org_id = user.get("org_id")
    update_data = {k: v for k, v in body.dict().items() if v is not None}

    # ISR override audit trail
    if body.isr_retenido is not None:
        current = supabase.table("planilla_empleados") \
            .select("isr_retenido").eq("id", empleado_id).eq("org_id", org_id).execute()
        if current.data:
            update_data["isr_override_original"] = current.data[0].get("isr_retenido", 0)
            update_data["isr_override_by"] = user.get("user_id")
            update_data["isr_override_at"] = datetime.now(timezone.utc).isoformat()

    result = supabase.table("planilla_empleados") \
        .update(update_data) \
        .eq("id", empleado_id) \
        .eq("org_id", org_id) \
        .execute()

    _recalc_planilla_totals(supabase, planilla_id)
    return {"empleado": result.data[0] if result.data else None}


@router.post("/{planilla_id}/add-empleado")
async def agregar_empleado(
    planilla_id: str,
    body: EmpleadoManual,
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    org_id = user.get("org_id")

    calc = _calcular_empleado({
        "salario_base": body.salario_base,
        "bonificaciones": body.bonificaciones,
        "horas_extras": body.horas_extras,
        "comisiones": body.comisiones,
        "isr_override": body.isr_override,
    }, supabase)

    result = supabase.table("planilla_empleados").insert({
        "planilla_id": planilla_id,
        "org_id": org_id,
        "nombre": body.nombre,
        "nit": body.nit,
        "dui": body.dui,
        "codigo_ingreso": body.codigo_ingreso,
        "salario_base": body.salario_base,
        "bonificaciones": body.bonificaciones,
        "horas_extras": body.horas_extras,
        "comisiones": body.comisiones,
        **calc,
    }).execute()

    _recalc_planilla_totals(supabase, planilla_id)
    return {"empleado": result.data[0] if result.data else None}


@router.post("/{planilla_id}/confirm")
async def confirmar_planilla(
    planilla_id: str,
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    """Confirmar planilla - genera lineas F14 automaticamente."""
    org_id = user.get("org_id")

    planilla = supabase.table("planilla_resumen") \
        .select("*").eq("id", planilla_id).eq("org_id", org_id).execute()
    if not planilla.data:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")
    if planilla.data[0]["status"] != "draft":
        raise HTTPException(status_code=400, detail="Planilla ya confirmada")

    periodo = planilla.data[0]["periodo"]

    per_check = supabase.table("renta_periodos") \
        .select("id").eq("org_id", org_id).eq("periodo", periodo).execute()
    if not per_check.data:
        supabase.table("renta_periodos").insert({
            "org_id": org_id, "periodo": periodo
        }).execute()

    empleados = supabase.table("planilla_empleados") \
        .select("*").eq("planilla_id", planilla_id).execute()

    created = 0
    for emp in (empleados.data or []):
        supabase.table("renta_retenciones").insert({
            "org_id": org_id,
            "periodo": periodo,
            "origen": "planilla",
            "planilla_id": planilla_id,
            "domicilio": 1,
            "codigo_pais": "9300",
            "nombre_razon": emp["nombre"],
            "nit_nif": emp.get("nit"),
            "dui": emp.get("dui"),
            "codigo_ingreso": emp.get("codigo_ingreso", "01"),
            "monto_devengado": emp.get("total_devengado", 0),
            "monto_bonificaciones": emp.get("bonificaciones", 0),
            "impuesto_retenido": emp.get("isr_retenido", 0),
            "aguinaldo_exento": emp.get("aguinaldo_exento", 0),
            "aguinaldo_gravado": emp.get("aguinaldo_gravado", 0),
            "afp": emp.get("afp", 0),
            "isss": emp.get("isss", 0),
            "inpep": emp.get("inpep", 0),
            "ipsfa": emp.get("ipsfa", 0),
            "cefafa": emp.get("cefafa", 0),
            "bienestar_mag": emp.get("bienestar_mag", 0),
            "isss_ivm": emp.get("isss_ivm", 0),
            "tipo_operacion": emp.get("tipo_operacion", 1),
            "clasificacion": emp.get("clasificacion", 2),
            "sector": emp.get("sector", 4),
            "tipo_costo_gasto": emp.get("tipo_costo_gasto", 7),
        }).execute()
        created += 1

    supabase.table("planilla_resumen") \
        .update({"status": "confirmed", "updated_at": datetime.now(timezone.utc).isoformat()}) \
        .eq("id", planilla_id).execute()

    from app.routers.renta_router import _update_periodo_totals
    _update_periodo_totals(supabase, org_id, periodo)

    # Generate accounting entry (non-blocking — if it fails, planilla is still confirmed)
    entry_id = None
    try:
        from app.services.contabilidad_service import generate_planilla_entry
        planilla_totals = supabase.table("planilla_resumen") \
            .select("*").eq("id", planilla_id).execute()
        if planilla_totals.data:
            entry_id = await generate_planilla_entry(
                supabase, org_id, planilla_totals.data[0], user.get("user_id")
            )
            if entry_id:
                logger.info(f"Partida contable {entry_id} generada para planilla {planilla_id}")
    except Exception as e:
        logger.warning(f"Partida contable no generada (no bloquea): {e}")

    return {
        "confirmed": True,
        "empleados_to_f14": created,
        "periodo": periodo,
        "journal_entry_id": entry_id,
    }


@router.delete("/{planilla_id}")
async def eliminar_planilla(
    planilla_id: str,
    supabase=Depends(get_supabase),
    user=Depends(get_current_user),
):
    org_id = user.get("org_id")

    planilla = supabase.table("planilla_resumen") \
        .select("status").eq("id", planilla_id).eq("org_id", org_id).execute()
    if planilla.data and planilla.data[0]["status"] != "draft":
        raise HTTPException(status_code=400, detail="Solo se pueden eliminar planillas en borrador")

    supabase.table("planilla_resumen") \
        .delete().eq("id", planilla_id).eq("org_id", org_id).execute()

    return {"deleted": True}
