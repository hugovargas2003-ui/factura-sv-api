"""
export_service.py — DTE history export in XLSX and PDF formats.

Location: app/services/export_service.py
Dependencies: openpyxl (XLSX), fpdf2 (PDF) — both already available

⚠️ NEW FILE — does not modify any existing infrastructure.
"""

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF


# ---------------------------------------------------------------------------
# Type labels
# ---------------------------------------------------------------------------

TIPO_DTE_LABELS = {
    "01": "Factura",
    "03": "CCF",
    "04": "Nota de Remisión",
    "05": "Nota de Crédito",
    "06": "Nota de Débito",
    "07": "Comp. Retención",
    "08": "Liquidación",
    "09": "DCD",
    "11": "FEXE",
    "14": "FSE",
    "15": "CD",
}


# ---------------------------------------------------------------------------
# Query helper
# ---------------------------------------------------------------------------

async def fetch_dtes_for_export(
    supabase_client: Any,
    org_id: str,
    date_from: str | None = None,
    date_to: str | None = None,
    tipo_dte: str | None = None,
    estado: str | None = None,
) -> list[dict]:
    """
    Fetch DTE rows for export.  Uses denormalized columns only.
    """
    query = (
        supabase_client.table("dtes")
        .select(
            "fecha_emision, tipo_dte, numero_control, codigo_generacion, "
            "receptor_nombre, receptor_nit, monto_total, iva, "
            "total_gravada, total_exenta, sello_recibido, estado"
        )
        .eq("org_id", org_id)
        .order("fecha_emision", desc=True)
    )

    if date_from:
        query = query.gte("fecha_emision", date_from)
    if date_to:
        query = query.lte("fecha_emision", date_to)
    if tipo_dte:
        query = query.eq("tipo_dte", tipo_dte)
    if estado:
        query = query.eq("estado", estado)

    resp = query.execute()
    return resp.data or []


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

HEADERS = [
    "Fecha",
    "Tipo",
    "Número Control",
    "Código Generación",
    "Receptor",
    "NIT Receptor",
    "Monto Total",
    "IVA",
    "Total Gravada",
    "Total Exenta",
    "Sello MH",
    "Estado",
]


def generate_xlsx(rows: list[dict], emisor_name: str = "FACTURA-SV") -> bytes:
    """Generate Excel workbook from DTE rows.  Returns bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DTEs"

    # -- Styles --
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_format = '#,##0.00'
    date_generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    # -- Title row --
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Reporte de DTEs — {emisor_name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"Generado: {date_generated}"
    subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center")

    # -- Headers (row 4) --
    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # -- Data rows --
    for row_idx, dte in enumerate(rows, start=5):
        tipo_label = TIPO_DTE_LABELS.get(dte.get("tipo_dte", ""), dte.get("tipo_dte", ""))
        data = [
            dte.get("fecha_emision", ""),
            tipo_label,
            dte.get("numero_control", ""),
            dte.get("codigo_generacion", ""),
            dte.get("receptor_nombre", ""),
            dte.get("receptor_nit", ""),
            dte.get("monto_total", 0),
            dte.get("iva", 0),
            dte.get("total_gravada", 0),
            dte.get("total_exenta", 0),
            dte.get("sello_recibido", ""),
            dte.get("estado", ""),
        ]
        for col_num, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # Money formatting for columns 7-10
            if col_num in (7, 8, 9, 10) and isinstance(val, (int, float)):
                cell.number_format = money_format

        # Conditional color for Estado
        estado_cell = ws.cell(row=row_idx, column=12)
        if estado_cell.value == "PROCESADO":
            estado_cell.font = Font(color="006B3C", bold=True)
        elif estado_cell.value == "RECHAZADO":
            estado_cell.font = Font(color="CC0000", bold=True)

    # -- Summary row --
    summary_row = len(rows) + 5
    ws.cell(row=summary_row, column=5, value="TOTALES:").font = Font(bold=True)
    for col_num in (7, 8, 9, 10):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        cell = ws.cell(
            row=summary_row,
            column=col_num,
            value=f"=SUM({col_letter}5:{col_letter}{summary_row - 1})",
        )
        cell.number_format = money_format
        cell.font = Font(bold=True)
        cell.border = thin_border

    # -- Column widths --
    widths = [12, 10, 36, 38, 30, 16, 14, 12, 14, 14, 44, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze pane
    ws.freeze_panes = "A5"

    # Auto-filter
    ws.auto_filter.ref = f"A4:L{len(rows) + 4}"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

class DTEReportPDF(FPDF):
    """Landscape PDF table report of DTEs."""

    def __init__(self, emisor_name: str = "FACTURA-SV"):
        super().__init__(orientation="L", unit="mm", format="letter")
        self.emisor_name = emisor_name
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 8, f"Reporte de DTEs - {self.emisor_name}", ln=True, align="C")
        self.set_font("Helvetica", "", 8)
        self.cell(
            0, 5,
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            ln=True, align="C",
        )
        self.ln(3)
        self._draw_table_header()

    def _draw_table_header(self):
        self.set_font("Helvetica", "B", 7)
        self.set_fill_color(31, 78, 121)
        self.set_text_color(255, 255, 255)
        cols = self._col_widths()
        headers = ["Fecha", "Tipo", "Número Control", "Receptor", "Monto", "IVA", "Sello MH", "Estado"]
        for w, h in zip(cols, headers):
            self.cell(w, 6, h, border=1, fill=True, align="C")
        self.ln()
        self.set_text_color(0, 0, 0)

    @staticmethod
    def _col_widths() -> list[int]:
        # Total ≈ 259mm (letter landscape minus margins)
        return [22, 18, 52, 50, 22, 18, 55, 22]

    def add_dte_row(self, dte: dict):
        cols = self._col_widths()
        self.set_font("Helvetica", "", 7)
        tipo_label = TIPO_DTE_LABELS.get(dte.get("tipo_dte", ""), dte.get("tipo_dte", ""))
        monto = f"${dte.get('monto_total', 0):,.2f}"
        iva = f"${dte.get('iva', 0):,.2f}"
        sello = (dte.get("sello_recibido") or "")[:40]

        estado = dte.get("estado", "")
        if estado == "PROCESADO":
            self.set_text_color(0, 107, 60)
        elif estado == "RECHAZADO":
            self.set_text_color(204, 0, 0)

        values = [
            dte.get("fecha_emision", ""),
            tipo_label,
            dte.get("numero_control", ""),
            (dte.get("receptor_nombre") or "")[:35],
            monto,
            iva,
            sello,
            estado,
        ]
        for w, v in zip(cols, values):
            self.cell(w, 5, str(v), border=1)
        self.ln()
        self.set_text_color(0, 0, 0)

    def add_summary(self, rows: list[dict]):
        self.ln(2)
        self.set_font("Helvetica", "B", 9)
        total_monto = sum(r.get("monto_total", 0) or 0 for r in rows)
        total_iva = sum(r.get("iva", 0) or 0 for r in rows)
        self.cell(0, 6, f"Total DTEs: {len(rows)}   |   Monto Total: ${total_monto:,.2f}   |   IVA Total: ${total_iva:,.2f}", ln=True)


def generate_pdf(rows: list[dict], emisor_name: str = "FACTURA-SV") -> bytes:
    """Generate PDF report from DTE rows.  Returns bytes."""
    pdf = DTEReportPDF(emisor_name=emisor_name)
    pdf.add_page()

    for dte in rows:
        pdf.add_dte_row(dte)

    pdf.add_summary(rows)

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.getvalue()
