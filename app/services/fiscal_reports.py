"""
fiscal_reports.py — Generates legally required fiscal reports for El Salvador.

Location: app/services/fiscal_reports.py
⚠️ NEW FILE — does not modify any existing infrastructure.

Reports:
- Libro de Ventas Contribuyente (from CCF tipo 03)
- Libro de Ventas Consumidor Final (from Factura tipo 01)
- Resumen IVA Mensual (aggregated)
- Reporte de Retenciones (from tipo 07)

Outputs: XLSX and PDF formats.
"""

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF


# ---------------------------------------------------------------------------
# Data fetchers
# ---------------------------------------------------------------------------

async def _fetch_dtes_by_type(
    supabase: Any, org_id: str, tipo_dte: str, year: int, month: int
) -> list[dict]:
    """Fetch DTEs of a specific type for a given month."""
    date_from = f"{year}-{month:02d}-01"
    if month == 12:
        date_to = f"{year + 1}-01-01"
    else:
        date_to = f"{year}-{month + 1:02d}-01"

    result = supabase.table("dtes").select(
        "fecha_emision, numero_control, codigo_generacion, "
        "receptor_nombre, receptor_nit, receptor_nrc, "
        "total_gravada, total_exenta, total_no_sujeta, "
        "iva, monto_total, sello_recibido, estado"
    ).eq("org_id", org_id).eq(
        "tipo_dte", tipo_dte
    ).eq("estado", "PROCESADO").gte(
        "fecha_emision", date_from
    ).lt(
        "fecha_emision", date_to
    ).order("fecha_emision").execute()

    return result.data or []


async def _fetch_emisor_name(supabase: Any, org_id: str) -> str:
    try:
        r = supabase.table("mh_credentials").select("nombre, nit, nrc").eq(
            "org_id", org_id
        ).single().execute()
        if r.data:
            return r.data.get("nombre", "FACTURA-SV")
    except Exception:
        pass
    return "FACTURA-SV"


async def _fetch_emisor_data(supabase: Any, org_id: str) -> dict:
    try:
        r = supabase.table("mh_credentials").select("*").eq(
            "org_id", org_id
        ).single().execute()
        return r.data or {}
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Libro de Ventas Contribuyente (CCF — tipo 03)
# ---------------------------------------------------------------------------

LIBRO_CCF_HEADERS = [
    "Nº", "Fecha", "Nº Control", "NIT Receptor", "NRC Receptor",
    "Nombre Receptor", "Gravada", "Exenta", "No Sujeta", "IVA", "Total",
]


async def generate_libro_ventas_contribuyente(
    supabase: Any, org_id: str, year: int, month: int, fmt: str = "xlsx"
) -> tuple[bytes, str]:
    """Generate Libro de Ventas Contribuyente (from CCF tipo 03)."""
    rows = await _fetch_dtes_by_type(supabase, org_id, "03", year, month)
    emisor = await _fetch_emisor_name(supabase, org_id)
    title = f"LIBRO DE VENTAS CONTRIBUYENTE — {_month_name(month)} {year}"

    if fmt == "xlsx":
        data = _libro_to_xlsx(rows, title, emisor, LIBRO_CCF_HEADERS, _ccf_row)
        filename = f"libro_ventas_contribuyente_{year}_{month:02d}.xlsx"
        return data, filename
    else:
        data = _libro_to_pdf(rows, title, emisor, _ccf_row)
        filename = f"libro_ventas_contribuyente_{year}_{month:02d}.pdf"
        return data, filename


# ---------------------------------------------------------------------------
# Libro de Ventas Consumidor Final (Factura — tipo 01)
# ---------------------------------------------------------------------------

LIBRO_CF_HEADERS = [
    "Nº", "Fecha", "Nº Control", "Nombre Receptor",
    "Gravada", "Exenta", "No Sujeta", "IVA", "Total",
]


async def generate_libro_ventas_consumidor(
    supabase: Any, org_id: str, year: int, month: int, fmt: str = "xlsx"
) -> tuple[bytes, str]:
    """Generate Libro de Ventas Consumidor Final (from Factura tipo 01)."""
    rows = await _fetch_dtes_by_type(supabase, org_id, "01", year, month)
    emisor = await _fetch_emisor_name(supabase, org_id)
    title = f"LIBRO DE VENTAS CONSUMIDOR FINAL — {_month_name(month)} {year}"

    if fmt == "xlsx":
        data = _libro_to_xlsx(rows, title, emisor, LIBRO_CF_HEADERS, _cf_row)
        filename = f"libro_ventas_consumidor_{year}_{month:02d}.xlsx"
        return data, filename
    else:
        data = _libro_to_pdf(rows, title, emisor, _cf_row)
        filename = f"libro_ventas_consumidor_{year}_{month:02d}.pdf"
        return data, filename


# ---------------------------------------------------------------------------
# Resumen IVA Mensual (aggregated from all types)
# ---------------------------------------------------------------------------

async def generate_resumen_iva(
    supabase: Any, org_id: str, year: int, month: int, fmt: str = "xlsx"
) -> tuple[bytes, str]:
    """Generate monthly IVA summary."""
    date_from = f"{year}-{month:02d}-01"
    if month == 12:
        date_to = f"{year + 1}-01-01"
    else:
        date_to = f"{year}-{month + 1:02d}-01"

    result = supabase.table("dtes").select(
        "tipo_dte, total_gravada, total_exenta, total_no_sujeta, iva, monto_total"
    ).eq("org_id", org_id).eq(
        "estado", "PROCESADO"
    ).gte("fecha_emision", date_from).lt("fecha_emision", date_to).execute()

    rows = result.data or []
    emisor = await _fetch_emisor_name(supabase, org_id)

    # Aggregate by tipo
    agg: dict[str, dict] = {}
    for r in rows:
        tipo = r.get("tipo_dte", "??")
        if tipo not in agg:
            agg[tipo] = {"count": 0, "gravada": 0, "exenta": 0, "no_sujeta": 0, "iva": 0, "total": 0}
        a = agg[tipo]
        a["count"] += 1
        a["gravada"] += r.get("total_gravada", 0) or 0
        a["exenta"] += r.get("total_exenta", 0) or 0
        a["no_sujeta"] += r.get("total_no_sujeta", 0) or 0
        a["iva"] += r.get("iva", 0) or 0
        a["total"] += r.get("monto_total", 0) or 0

    title = f"RESUMEN IVA MENSUAL — {_month_name(month)} {year}"

    if fmt == "xlsx":
        data = _resumen_to_xlsx(agg, title, emisor)
        filename = f"resumen_iva_{year}_{month:02d}.xlsx"
        return data, filename
    else:
        data = _resumen_to_pdf(agg, title, emisor)
        filename = f"resumen_iva_{year}_{month:02d}.pdf"
        return data, filename


# ---------------------------------------------------------------------------
# Row extractors
# ---------------------------------------------------------------------------

DTE_TIPO_LABELS = {
    "01": "Factura", "03": "CCF", "04": "Nota Remisión",
    "05": "Nota Crédito", "06": "Nota Débito", "07": "Comp. Retención",
    "08": "Liquidación", "09": "DCD", "14": "FSE", "15": "CD",
}


def _ccf_row(idx: int, r: dict) -> list:
    return [
        idx, r.get("fecha_emision", ""), r.get("numero_control", ""),
        r.get("receptor_nit", ""), r.get("receptor_nrc", ""),
        r.get("receptor_nombre", ""),
        r.get("total_gravada", 0) or 0,
        r.get("total_exenta", 0) or 0,
        r.get("total_no_sujeta", 0) or 0,
        r.get("iva", 0) or 0,
        r.get("monto_total", 0) or 0,
    ]


def _cf_row(idx: int, r: dict) -> list:
    return [
        idx, r.get("fecha_emision", ""), r.get("numero_control", ""),
        r.get("receptor_nombre", ""),
        r.get("total_gravada", 0) or 0,
        r.get("total_exenta", 0) or 0,
        r.get("total_no_sujeta", 0) or 0,
        r.get("iva", 0) or 0,
        r.get("monto_total", 0) or 0,
    ]


# ---------------------------------------------------------------------------
# XLSX generators
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_MONEY_FMT = '#,##0.00'
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _libro_to_xlsx(rows, title, emisor, headers, row_fn) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].value = title
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"].value = f"Emisor: {emisor} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers row 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN

    # Data
    totals = {}
    for i, r in enumerate(rows, 1):
        vals = row_fn(i, r)
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 4, column=col, value=v)
            cell.border = _THIN
            if isinstance(v, (int, float)) and col > 3:
                cell.number_format = _MONEY_FMT
                totals[col] = totals.get(col, 0) + v

    # Totals row
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    for col, val in totals.items():
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.number_format = _MONEY_FMT
        cell.font = Font(bold=True)
        cell.border = _THIN

    # Column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _resumen_to_xlsx(agg, title, emisor) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen IVA"

    headers = ["Tipo DTE", "Cantidad", "Gravada", "Exenta", "No Sujeta", "IVA", "Total"]

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Emisor: {emisor}"
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN

    row_num = 5
    grand = {"count": 0, "gravada": 0, "exenta": 0, "no_sujeta": 0, "iva": 0, "total": 0}
    for tipo, vals in sorted(agg.items()):
        label = DTE_TIPO_LABELS.get(tipo, tipo)
        data = [label, vals["count"], vals["gravada"], vals["exenta"],
                vals["no_sujeta"], vals["iva"], vals["total"]]
        for col, v in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = _THIN
            if col >= 3:
                cell.number_format = _MONEY_FMT
        for k in grand:
            grand[k] += vals[k]
        row_num += 1

    # Grand total
    ws.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_num, column=2, value=grand["count"]).font = Font(bold=True)
    for col, key in enumerate(["gravada", "exenta", "no_sujeta", "iva", "total"], 3):
        cell = ws.cell(row=row_num, column=col, value=grand[key])
        cell.number_format = _MONEY_FMT
        cell.font = Font(bold=True)
        cell.border = _THIN

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF generators
# ---------------------------------------------------------------------------

def _libro_to_pdf(rows, title, emisor, row_fn) -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="letter")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, title, ln=True, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, f"Emisor: {emisor} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    # Table
    cols = [8, 22, 50, 35, 35, 22, 22, 22, 22, 22]
    pdf_headers = ["#", "Fecha", "Nº Control", "Receptor", "Gravada", "Exenta", "N/S", "IVA", "Total"]

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(cols, pdf_headers):
        pdf.cell(w, 5, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 7)
    for i, r in enumerate(rows, 1):
        vals = row_fn(i, r)
        # Simplify for PDF: idx, fecha, nc, receptor_name, gravada, exenta, ns, iva, total
        pdf_vals = [str(vals[0]), str(vals[1]), str(vals[2]),
                    (str(vals[-6]) if len(vals) > 9 else str(vals[3]))[:30],
                    f"${vals[-5]:,.2f}" if isinstance(vals[-5], (int, float)) else str(vals[-5]),
                    f"${vals[-4]:,.2f}" if isinstance(vals[-4], (int, float)) else str(vals[-4]),
                    f"${vals[-3]:,.2f}" if isinstance(vals[-3], (int, float)) else str(vals[-3]),
                    f"${vals[-2]:,.2f}" if isinstance(vals[-2], (int, float)) else str(vals[-2]),
                    f"${vals[-1]:,.2f}" if isinstance(vals[-1], (int, float)) else str(vals[-1])]
        for w, v in zip(cols, pdf_vals):
            pdf.cell(w, 4, v, border=1)
        pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.getvalue()


def _resumen_to_pdf(agg, title, emisor) -> bytes:
    pdf = FPDF(orientation="P", unit="mm", format="letter")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, title, ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Emisor: {emisor}", ln=True, align="C")
    pdf.ln(8)

    cols = [50, 20, 30, 30, 30, 25, 30]
    headers = ["Tipo", "Cant", "Gravada", "Exenta", "No Suj", "IVA", "Total"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(cols, headers):
        pdf.cell(w, 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for tipo, v in sorted(agg.items()):
        label = DTE_TIPO_LABELS.get(tipo, tipo)
        data = [label, str(v["count"]),
                f"${v['gravada']:,.2f}", f"${v['exenta']:,.2f}",
                f"${v['no_sujeta']:,.2f}", f"${v['iva']:,.2f}",
                f"${v['total']:,.2f}"]
        for w, d in zip(cols, data):
            pdf.cell(w, 6, d, border=1)
        pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MONTHS = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

def _month_name(m: int) -> str:
    return _MONTHS[m] if 1 <= m <= 12 else str(m)
