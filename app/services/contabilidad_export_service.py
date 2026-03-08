"""
Contabilidad Export Service — Genera XLSX profesionales.
Libro Diario, Libro Mayor, Estado de Resultados.
Usa contabilidad_service.py existente (no lo modifica).
"""
import logging
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F2937")
SUBTITLE_FONT = Font(name="Arial", size=11, color="6B7280")
DATA_FONT = Font(name="Arial", size=10)
MONEY_FONT = Font(name="Arial", size=10)
TOTAL_FONT = Font(name="Arial", size=11, bold=True)
TOTAL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
MONEY_FORMAT = '#,##0.00'


def _write_company_header(ws, org_name: str, org_nit: str, report_title: str, periodo: str, start_row: int = 1, num_cols: int = 6):
    """Write company header block at top of report."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
    cell = ws.cell(row=start_row, column=1, value=org_name or "FACTURA-SV")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=num_cols)
    cell = ws.cell(row=start_row + 1, column=1, value=f"NIT: {org_nit}" if org_nit else "")
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=num_cols)
    cell = ws.cell(row=start_row + 2, column=1, value=report_title)
    cell.font = Font(name="Arial", size=12, bold=True, color="4F46E5")
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=num_cols)
    cell = ws.cell(row=start_row + 3, column=1, value=f"Período: {periodo}")
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="center")

    return start_row + 5  # Next available row


def _style_header_row(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def generate_libro_diario_xlsx(entries: list, org_name: str, org_nit: str, periodo: str) -> bytes:
    """
    Genera Libro Diario en XLSX con formato profesional.
    entries = list of journal_entries con journal_entry_lines nested.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    # Column widths
    col_widths = [12, 10, 15, 30, 15, 15]
    headers = ["Fecha", "Partida #", "Código", "Cuenta / Concepto", "Debe", "Haber"]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Company header
    data_row = _write_company_header(ws, org_name, org_nit, "LIBRO DIARIO", periodo, num_cols=6)

    # Table headers
    for i, h in enumerate(headers, 1):
        ws.cell(row=data_row, column=i, value=h)
    _style_header_row(ws, data_row, 6)
    data_row += 1

    grand_debe = 0
    grand_haber = 0

    for entry in entries:
        # Entry header row
        fecha = entry.get("fecha", "")
        numero = entry.get("numero", "")
        descripcion = entry.get("descripcion", "")

        ws.cell(row=data_row, column=1, value=fecha).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=data_row, column=2, value=f"#{numero}").font = Font(name="Arial", size=10, bold=True)
        ws.merge_cells(start_row=data_row, start_column=3, end_row=data_row, end_column=4)
        ws.cell(row=data_row, column=3, value=descripcion).font = Font(name="Arial", size=10, italic=True, color="6B7280")
        for col in range(1, 7):
            ws.cell(row=data_row, column=col).border = THIN_BORDER
        data_row += 1

        # Entry lines
        lines = entry.get("journal_entry_lines", entry.get("lines", []))
        entry_debe = 0
        entry_haber = 0

        for line in lines:
            codigo = line.get("cuenta_codigo", "")
            cuenta = line.get("cuenta_nombre", "")
            debe = float(line.get("debe", 0))
            haber = float(line.get("haber", 0))

            ws.cell(row=data_row, column=3, value=codigo).font = Font(name="Arial", size=10, color="4F46E5")
            ws.cell(row=data_row, column=4, value=cuenta).font = DATA_FONT

            if debe > 0:
                c = ws.cell(row=data_row, column=5, value=debe)
                c.font = MONEY_FONT
                c.number_format = MONEY_FORMAT
            if haber > 0:
                c = ws.cell(row=data_row, column=6, value=haber)
                c.font = MONEY_FONT
                c.number_format = MONEY_FORMAT
                # Indent haber accounts
                ws.cell(row=data_row, column=4).alignment = Alignment(indent=2)

            for col in range(1, 7):
                ws.cell(row=data_row, column=col).border = THIN_BORDER

            entry_debe += debe
            entry_haber += haber
            data_row += 1

        # Entry subtotal
        ws.cell(row=data_row, column=4, value="Subtotal partida").font = Font(name="Arial", size=10, bold=True, color="6B7280")
        c = ws.cell(row=data_row, column=5, value=entry_debe)
        c.font = TOTAL_FONT
        c.number_format = MONEY_FORMAT
        c = ws.cell(row=data_row, column=6, value=entry_haber)
        c.font = TOTAL_FONT
        c.number_format = MONEY_FORMAT
        for col in range(1, 7):
            ws.cell(row=data_row, column=col).fill = TOTAL_FILL
            ws.cell(row=data_row, column=col).border = THIN_BORDER
        data_row += 1

        grand_debe += entry_debe
        grand_haber += entry_haber

        # Blank row between entries
        data_row += 1

    # Grand total
    data_row += 1
    ws.cell(row=data_row, column=4, value="TOTAL GENERAL").font = Font(name="Arial", size=12, bold=True)
    c = ws.cell(row=data_row, column=5, value=grand_debe)
    c.font = Font(name="Arial", size=12, bold=True, color="059669")
    c.number_format = MONEY_FORMAT
    c = ws.cell(row=data_row, column=6, value=grand_haber)
    c.font = Font(name="Arial", size=12, bold=True, color="059669")
    c.number_format = MONEY_FORMAT
    for col in range(1, 7):
        ws.cell(row=data_row, column=col).border = Border(top=Side(style="double", color="1F2937"), bottom=Side(style="double", color="1F2937"))

    # Footer
    data_row += 2
    ws.cell(row=data_row, column=1, value=f"Generado por FACTURA-SV — {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(name="Arial", size=8, color="9CA3AF")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_estado_resultados_xlsx(cuentas_balance: list, org_name: str, org_nit: str, periodo: str) -> bytes:
    """
    Genera Estado de Resultados en XLSX.
    cuentas_balance = output de get_balance_general()["cuentas"]
    Estructura: Ingresos (4xxx) - Costos (5xxx) - Gastos (6xxx) = Utilidad/Pérdida
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    col_widths = [15, 40, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    data_row = _write_company_header(ws, org_name, org_nit, "ESTADO DE RESULTADOS", periodo, num_cols=3)

    # Classify accounts
    ingresos = []
    costos = []
    gastos = []

    for c in cuentas_balance:
        codigo = str(c.get("cuenta_codigo", ""))
        if codigo.startswith("4"):
            # Income: natural credit balance (haber - debe)
            saldo = abs(float(c.get("haber", 0)) - float(c.get("debe", 0)))
            ingresos.append({**c, "saldo": saldo})
        elif codigo.startswith("5"):
            # Cost: natural debit balance (debe - haber)
            saldo = abs(float(c.get("debe", 0)) - float(c.get("haber", 0)))
            costos.append({**c, "saldo": saldo})
        elif codigo.startswith("6"):
            # Expense: natural debit balance (debe - haber)
            saldo = abs(float(c.get("debe", 0)) - float(c.get("haber", 0)))
            gastos.append({**c, "saldo": saldo})

    total_ingresos = sum(i["saldo"] for i in ingresos)
    total_costos = sum(i["saldo"] for i in costos)
    total_gastos = sum(i["saldo"] for i in gastos)
    utilidad = total_ingresos - total_costos - total_gastos

    def write_section(title, items, total, row, color):
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name="Arial", size=12, bold=True, color=color)
        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

        for item in items:
            ws.cell(row=row, column=1, value=item.get("cuenta_codigo", "")).font = Font(name="Arial", size=10, color="4F46E5")
            ws.cell(row=row, column=2, value=item.get("cuenta_nombre", "")).font = DATA_FONT
            c = ws.cell(row=row, column=3, value=item["saldo"])
            c.font = MONEY_FONT
            c.number_format = MONEY_FORMAT
            c.alignment = Alignment(horizontal="right")
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        # Section total
        ws.cell(row=row, column=2, value=f"Total {title}").font = TOTAL_FONT
        c = ws.cell(row=row, column=3, value=total)
        c.font = Font(name="Arial", size=11, bold=True, color=color)
        c.number_format = MONEY_FORMAT
        c.alignment = Alignment(horizontal="right")
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = TOTAL_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        return row + 1  # Blank row

    # Write sections
    data_row = write_section("INGRESOS", ingresos, total_ingresos, data_row, "059669")
    data_row = write_section("COSTOS DE VENTA", costos, total_costos, data_row, "DC2626")

    # Utilidad Bruta
    ws.cell(row=data_row, column=2, value="UTILIDAD BRUTA").font = Font(name="Arial", size=11, bold=True)
    c = ws.cell(row=data_row, column=3, value=total_ingresos - total_costos)
    c.font = Font(name="Arial", size=11, bold=True)
    c.number_format = MONEY_FORMAT
    c.alignment = Alignment(horizontal="right")
    for col in range(1, 4):
        ws.cell(row=data_row, column=col).border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    data_row += 2

    data_row = write_section("GASTOS DE OPERACIÓN", gastos, total_gastos, data_row, "D97706")

    # Utilidad/Pérdida Neta
    data_row += 1
    ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=2)
    label = "UTILIDAD NETA DEL PERÍODO" if utilidad >= 0 else "PÉRDIDA NETA DEL PERÍODO"
    color = "059669" if utilidad >= 0 else "DC2626"
    ws.cell(row=data_row, column=1, value=label).font = Font(name="Arial", size=14, bold=True, color=color)
    c = ws.cell(row=data_row, column=3, value=abs(utilidad))
    c.font = Font(name="Arial", size=14, bold=True, color=color)
    c.number_format = MONEY_FORMAT
    c.alignment = Alignment(horizontal="right")
    for col in range(1, 4):
        ws.cell(row=data_row, column=col).border = Border(top=Side(style="double", color="1F2937"), bottom=Side(style="double", color="1F2937"))

    # Footer
    data_row += 3
    ws.cell(row=data_row, column=1, value=f"Generado por FACTURA-SV — {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(name="Arial", size=8, color="9CA3AF")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
